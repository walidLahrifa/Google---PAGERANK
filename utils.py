# Creation Data    :    10/11/2020
# Last Modification:    10/11/2020
# Authors          :    'LAHRIFA Walid' & 'EN NIARI SAAD'
# Short Description:    Utils methods for : Generate Statistics / Export Statistics to Excel

from timeit import default_timer as timer
from networkx import erdos_renyi_graph
from pagerank import pagerank
from pagerankx import pagerankx
import openpyxl


def generate_statistics(N_start, N_end, step, p, zap_factor):
    """
    Generate statistics about the time execution of
    Parameters
    ----------
    N_start : Initial Number of Nodes
    N_end   : Max of Node
    step    : Used to increment the Node Number of Generated raph
    p       : Probability of edges creation between 0 & 1
    zap_factor : Used in Page Rank Algorithm, recommended between 0.1 & 0.2

    Returns
    -------
    statistics : List of number of nodes with associated time execution of Graph Generation & Page Rank Algorithm

    """
    statistics = []
    N = N_start

    while N < N_end:
        # Start Timer
        start_gen = timer()

        # construct the Graph
        graph = erdos_renyi_graph(n=N, p=p, seed=None,
                                      directed=True)

        end_gen = timer()

        start_rank = timer()

        # Calculate the Page Rank
        pagerankx(graph, zap_factor)



        end_rank = timer()

        statistics.append([N,len(graph.edges), end_gen - start_gen, end_rank - start_rank])
        print("------------------------------------------------------------------");
        print("Graph with Nodes number:    " + str(N))
        print("Number of edges        :    " + str(len(graph.edges)))
        print("Generation Time:            " + str(end_gen - start_gen) + "s")
        print("Page Rank calculation Time: " + str(end_rank - start_rank) + "s")

        export_stat_excel(stat_gen=statistics)

        N = N + step
    print("------------------------------------------------------------------");
    print("**********  Statistics saved on \"Statistics_Gen.xlsx\"  **********");
    print("------------------------------------------------------------------");

    return statistics


def export_stat_excel(stat_gen):
    """

    Parameters
    ----------
    stat_gen : List of generated statistics

    Returns
    -------
    file 'Statistics_Gen.xlsx' saved in project folder

    """
    xfile = openpyxl.load_workbook('Statistics.xlsx')
    sheet_1 = xfile['Evolution']

    row = 9
    for st in stat_gen:
        sheet_1.cell(row, column=1).value = st[0]
        sheet_1.cell(row, column=2).value = st[1]
        sheet_1.cell(row, column=3).value = float(st[2])
        sheet_1.cell(row, column=4).value = float(st[3])
        row = row + 1

    xfile.save('Statistics_Gen.xlsx')


if __name__ == '__main__':
    generate_statistics(N_start=1, N_end=25001, step=1000, p=0.05, zap_factor=0.15)
